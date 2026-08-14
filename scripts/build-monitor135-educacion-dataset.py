"""
Educación — dimensión nueva de Monitor 135.

Consolida, para los 135 municipios de la Provincia de Buenos Aires, los
indicadores educativos prioritarios del mapa a partir de
`Mega_Base_Alsina_Educacion_PBA_v1.xlsx`:

1. RESUMEN_135 — indicadores consolidados por municipio (matrícula,
   trayectorias secundarias, composición estatal/privada, expansión).
2. EVOL_MUN_2015_2025 — serie 2015-2025 de unidades de servicio y
   matrícula, necesaria para la variación 2015-2025 y la brecha
   matrícula/red (RESUMEN_135 sólo trae 2019-2025 para nivel secundario).

El join contra el sistema NO usa el `municipio_id` del Excel (es un
código numérico tipo DINIECE, sin relación con el `PBA-XXX` del sitio):
se normaliza `municipio_nombre` con la misma función `norm()` de
scripts/build-monitor135-dataset.py y se valida contra las 135 claves
canónicas de Info_Municipios_nueva/pba_municipios_master_v2.csv.

No inventa valores. Un denominador ausente o cero da `None`, nunca 0.
Los porcentajes ya vienen expresados en puntos porcentuales en el Excel
(ej. 5.26 = 5,26%); acá se guardan como fracción (0.0526) para ser
consistentes con `fmtPct()` en monitor135-app.js, que multiplica por 100.

Anomalías reales detectadas y NO corregidas (ver docstring de cada
chequeo más abajo): abandono interanual negativo (retorno > abandono
neto, metodológicamente posible), variaciones % sobre una base inicial
chica, y valores extremos en alumnos por sede/sección — todas quedan
marcadas con `quality_flag`, nunca ocultas ni "corregidas".

Salida: assets/data/monitor135-educacion.json
"""
import csv
import json
import statistics
import unicodedata
import openpyxl

XLSX = 'Mega_Base_Alsina_Educacion_PBA_v1.xlsx'
MASTER_CSV = 'Info_Municipios_nueva/pba_municipios_master_v2.csv'
OUT_JSON = 'assets/data/monitor135-educacion.json'

FUENTE = 'DGCyE PBA — Mega Base Educación Alsina'

# El Excel nombra distinto a 4 municipios respecto de la clave canónica del
# sitio (Info_Municipios_nueva/pba_municipios_master_v2.csv). Alias
# documentado, no adivinado: verificado 1 a 1 contra el maestro.
ALIAS_EXCEL_TO_CANON = {
    '25 DE MAYO': 'VEINTICINCO DE MAYO',
    '9 DE JULIO': 'NUEVE DE JULIO',
    'CORONEL DE MARINA LEONARDO ROSALES': 'CORONEL ROSALES',
    'GENERAL JUAN MADARIAGA': 'GENERAL MADARIAGA',
    # sólo en EVOL_MUN_2015_2025, que en esta Mega Base trae el nombre sin el
    # prenombre de pila que sí usan RESUMEN_135 y el maestro del sitio.
    'GONZALES CHAVES': 'ADOLFO GONZALES CHAVES',
    'JUAREZ': 'BENITO JUAREZ',
}


def norm(s):
    # Misma normalización que scripts/build-monitor135-dataset.py: la Ñ es
    # una letra propia, no una vocal acentuada, y no debe perderse en la
    # descomposición NFKD (si no, "CAÑUELAS" pasaría a "CANUELAS" y dejaría
    # de matchear la clave que usan PATHS/GEO_DATA/MUNIS).
    s = s.replace('ñ', '\x00').replace('Ñ', '\x00')
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii').upper().strip()
    return s.replace('\x00', 'Ñ')


def f(v):
    """Convierte '', None en None. No inventa un valor. No trata 0 como
    faltante: en educación (matrícula, secciones, unidades) un 0 real es
    válido y distinto de "sin dato"."""
    if v is None or v == '':
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def pct_frac(v):
    """RESUMEN_135 guarda los porcentajes ya multiplicados por 100
    (ej. 5.26 = 5,26%). Se convierte a fracción para que fmtPct() del
    front (que multiplica por 100) siga funcionando igual que con las
    demás dimensiones del sitio."""
    v = f(v)
    return None if v is None else v / 100


def ratio(numer, denom):
    n, d = f(numer), f(denom)
    if n is None or d is None or d == 0:
        return None
    return n / d


def var_pct(v0, v1):
    v0, v1 = f(v0), f(v1)
    if v0 is None or v1 is None or v0 == 0:
        return None
    return (v1 - v0) / v0


# ── 1) claves canónicas del sitio (135) ──────────────────────────────
with open(MASTER_CSV, encoding='utf-8-sig') as fh:
    master_rows = list(csv.DictReader(fh))
assert len(master_rows) == 135, len(master_rows)
canon = {}  # clave normalizada -> {id, municipio, slug}
for r in master_rows:
    key = norm(r['municipio_nombre_fuente'])
    canon[key] = {'id': r['municipio_id_dataset'], 'municipio': r['municipio_nombre'], 'municipio_slug': r['municipio_slug']}
assert len(canon) == 135

# ── 2) RESUMEN_135 ────────────────────────────────────────────────────
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb['RESUMEN_135']
header = [c.value for c in ws[1]]
col = {h: i for i, h in enumerate(header)}
resumen = {}  # clave normalizada -> fila (dict de valores crudos)
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[col['municipio_nombre']] is None:
        continue
    key = ALIAS_EXCEL_TO_CANON.get(norm(str(row[col['municipio_nombre']])), norm(str(row[col['municipio_nombre']])))
    resumen[key] = {h: row[i] for h, i in col.items()}
assert len(resumen) == 135, f'RESUMEN_135: se esperaban 135 municipios, se encontraron {len(resumen)}'

unmatched = sorted(set(resumen) - set(canon))
missing_from_excel = sorted(set(canon) - set(resumen))
assert not unmatched, f'RESUMEN_135 tiene claves que no matchean el maestro del sitio: {unmatched}'
assert not missing_from_excel, f'Faltan en RESUMEN_135 municipios del maestro del sitio: {missing_from_excel}'

# ── 3) EVOL_MUN_2015_2025 (serie 2015-2025, unidades y matrícula) ────
ws2 = wb['EVOL_MUN_2015_2025']
header2 = [c.value for c in ws2[1]]
col2 = {h: i for i, h in enumerate(header2)}
evol = {}  # (clave, anio) -> {unidades_de_servicio, matricula}
for row in ws2.iter_rows(min_row=2, values_only=True):
    if row[col2['municipio_nombre']] is None:
        continue
    key = ALIAS_EXCEL_TO_CANON.get(norm(str(row[col2['municipio_nombre']])), norm(str(row[col2['municipio_nombre']])))
    anio = row[col2['anio']]
    evol[(key, anio)] = {
        'unidades_de_servicio': row[col2['unidades_de_servicio']],
        'matricula': row[col2['matricula']],
    }

for key in canon:
    assert (key, 2015) in evol and (key, 2025) in evol, f'EVOL_MUN_2015_2025: falta serie 2015/2025 para {key}'

# ── 4) indicadores por municipio ──────────────────────────────────────
raw = {}  # clave -> dict de valores ya calculados (antes de flags), para poder sacar percentiles/IQR después
for key in canon:
    r = resumen[key]
    e15 = evol[(key, 2015)]
    e25 = evol[(key, 2025)]

    matricula_2025 = f(r['alumnos_anuario_2025'])
    var_matricula_15_25 = var_pct(e15['matricula'], e25['matricula'])
    var_secundaria_19_25 = pct_frac(r['secundaria_variacion_2019_2025_pct'])
    alumnos_por_sede = ratio(r['alumnos_anuario_2025'], r['nomina_sedes_unicas_2025'])
    alumnos_por_seccion = ratio(r['alumnos_anuario_2025'], r['secciones_anuario_2025'])
    var_unidades_15_25 = var_pct(e15['unidades_de_servicio'], e25['unidades_de_servicio'])
    brecha = None if (var_matricula_15_25 is None or var_unidades_15_25 is None) else var_matricula_15_25 - var_unidades_15_25
    abandono_sec = pct_frac(r['sec_abandono_pct'])
    # promocion_sec/sobreedad_sec/participacion_estatal se registran en el
    # front como tipo:'seq' + unidad:'%' — ese combo espera el valor YA en
    # puntos porcentuales (ej. 90.91), no como fracción (a diferencia de
    # tipo:'div'/'div_alert', que sí esperan fracción porque fmtPct() la
    # multiplica por 100 — ver fmtValue() en monitor135-app.js). Mismo
    # criterio ya usado por el campo 'cud' de finanzas.
    promocion_sec = f(r['sec_promocion_efectiva_pct'])
    sobreedad_sec = f(r['sec_sobreedad_pct'])
    _part_estatal_frac = ratio(r['alumnos_estatal_2025'], r['alumnos_anuario_2025'])
    participacion_estatal = None if _part_estatal_frac is None else _part_estatal_frac * 100

    raw[key] = {
        'matricula_2025': matricula_2025,
        'var_matricula_15_25': var_matricula_15_25,
        'var_matricula_15_25_base': f(e15['matricula']),
        'var_matricula_15_25_final': f(e25['matricula']),
        'var_secundaria_19_25': var_secundaria_19_25,
        'var_secundaria_19_25_base': f(r['secundaria_matricula_2019']),
        'var_secundaria_19_25_final': f(r['secundaria_matricula_2025']),
        'alumnos_por_sede': alumnos_por_sede,
        'alumnos_por_seccion': alumnos_por_seccion,
        'var_unidades_15_25': var_unidades_15_25,
        'var_unidades_15_25_base': f(e15['unidades_de_servicio']),
        'var_unidades_15_25_final': f(e25['unidades_de_servicio']),
        'brecha': brecha,
        'abandono_sec': abandono_sec,
        'promocion_sec': promocion_sec,
        'sobreedad_sec': sobreedad_sec,
        'participacion_estatal': participacion_estatal,
        'sedes_2025': f(r['nomina_sedes_unicas_2025']),
        'unidades_servicio_2025': f(r['nomina_unidades_servicio_2025']),
        'establecimientos_2026': f(r['establecimientos_2026']),
        'actos_expansion_total': f(r['actos_expansion_total']),
        'ultimo_anio_expansion': f(r['ultimo_anio_expansion']),
    }

# ── 5) control de calidad ─────────────────────────────────────────────
# base_reducida: variación % calculada sobre una base inicial que está en
# el decil más bajo de la provincia. El corte se calcula (percentil 10),
# no se inventa un número redondo.
def p10(values):
    vals = sorted(v for v in values if v is not None and v > 0)
    if len(vals) < 10:
        return None
    return statistics.quantiles(vals, n=10)[0]

base_p10 = {
    'var_matricula_15_25': p10(r['var_matricula_15_25_base'] for r in raw.values()),
    'var_secundaria_19_25': p10(r['var_secundaria_19_25_base'] for r in raw.values()),
    'var_unidades_15_25': p10(r['var_unidades_15_25_base'] for r in raw.values()),
}


def iqr_bounds(values):
    vals = sorted(v for v in values if v is not None)
    if len(vals) < 10:
        return None
    q1, q3 = statistics.quantiles(vals, n=4)[0], statistics.quantiles(vals, n=4)[2]
    iqr = q3 - q1
    return (statistics.median(vals) - 3 * iqr, statistics.median(vals) + 3 * iqr)

extreme_bounds = {
    'alumnos_por_sede': iqr_bounds(r['alumnos_por_sede'] for r in raw.values()),
    'alumnos_por_seccion': iqr_bounds(r['alumnos_por_seccion'] for r in raw.values()),
}


def flag_base_reducida(field, base_value):
    thr = base_p10[field]
    return 'base_reducida' if (thr is not None and base_value is not None and base_value < thr) else None


def flag_extremo(field, value):
    bounds = extreme_bounds[field]
    if bounds is None or value is None:
        return None
    lo, hi = bounds
    return 'valor_extremo' if (value < lo or value > hi) else None


anomaly_counts = {'tasa_negativa': 0, 'base_reducida': 0, 'valor_extremo': 0}


def field(valor, periodo, fuente=FUENTE, url=None, quality_flag=None, **extra):
    if quality_flag:
        anomaly_counts[quality_flag] = anomaly_counts.get(quality_flag, 0) + 1
    out = {'valor': valor, 'periodo': periodo, 'fuente': fuente, 'estado': 'disponible' if valor is not None else 'no_disponible'}
    if url:
        out['url'] = url
    if quality_flag:
        out['quality_flag'] = quality_flag
    out.update(extra)
    return out


municipios = {}
for key, c in canon.items():
    r = raw[key]

    flag_vm = flag_base_reducida('var_matricula_15_25', r['var_matricula_15_25_base'])
    flag_vs = flag_base_reducida('var_secundaria_19_25', r['var_secundaria_19_25_base'])
    flag_vu = flag_base_reducida('var_unidades_15_25', r['var_unidades_15_25_base'])
    flag_ab = 'tasa_negativa' if (r['abandono_sec'] is not None and r['abandono_sec'] < 0) else None
    flag_aps = flag_extremo('alumnos_por_sede', r['alumnos_por_sede'])
    flag_apsec = flag_extremo('alumnos_por_seccion', r['alumnos_por_seccion'])

    municipios[key] = {
        'id': c['id'],
        'municipio': c['municipio'],
        'municipio_slug': c['municipio_slug'],
        'municipio_normalizado': key,

        'edu_matricula_total': field(r['matricula_2025'], '2025', fuente='DGCyE PBA — Anuario Estadístico 2025'),

        'edu_var_matricula_2015_2025': field(
            r['var_matricula_15_25'], '2015–2025', quality_flag=flag_vm,
            valor_inicial=r['var_matricula_15_25_base'], valor_final=r['var_matricula_15_25_final'],
            cambio_absoluto=None if (r['var_matricula_15_25_base'] is None or r['var_matricula_15_25_final'] is None) else r['var_matricula_15_25_final'] - r['var_matricula_15_25_base'],
        ),

        'edu_var_secundaria_2019_2025': field(
            r['var_secundaria_19_25'], '2019–2025', quality_flag=flag_vs,
            valor_inicial=r['var_secundaria_19_25_base'], valor_final=r['var_secundaria_19_25_final'],
            cambio_absoluto=None if (r['var_secundaria_19_25_base'] is None or r['var_secundaria_19_25_final'] is None) else r['var_secundaria_19_25_final'] - r['var_secundaria_19_25_base'],
        ),

        'edu_alumnos_por_sede': field(r['alumnos_por_sede'], '2025', quality_flag=flag_aps),
        'edu_alumnos_por_seccion': field(r['alumnos_por_seccion'], '2025', quality_flag=flag_apsec),

        'edu_var_unidades_2015_2025': field(
            r['var_unidades_15_25'], '2015–2025', quality_flag=flag_vu,
            valor_inicial=r['var_unidades_15_25_base'], valor_final=r['var_unidades_15_25_final'],
            cambio_absoluto=None if (r['var_unidades_15_25_base'] is None or r['var_unidades_15_25_final'] is None) else r['var_unidades_15_25_final'] - r['var_unidades_15_25_base'],
        ),

        'edu_brecha_matricula_red': field(
            r['brecha'], '2015–2025', fuente='Cálculo Alsina',
            componente_matricula=r['var_matricula_15_25'], componente_unidades=r['var_unidades_15_25'],
        ),

        'edu_abandono_secundario': field(
            r['abandono_sec'], '2024–2025', quality_flag=flag_ab,
            nota_metodologica='La tasa de abandono interanual puede ser negativa cuando el retorno de alumnos supera al abandono neto entre años. No se trata de un error de carga.' if flag_ab else None,
        ),
        'edu_promocion_secundaria': field(r['promocion_sec'], '2024–2025'),
        'edu_sobreedad_secundaria': field(r['sobreedad_sec'], '2025'),
        'edu_participacion_estatal': field(r['participacion_estatal'], '2025'),

        'edu_sedes_2025': field(r['sedes_2025'], '2025'),
        'edu_unidades_servicio_2025': field(r['unidades_servicio_2025'], '2025'),
        'edu_establecimientos_2026': field(r['establecimientos_2026'], '2026'),
        'edu_actos_expansion_total': field(r['actos_expansion_total'], '2015–2025 (fecha del acto)'),
        'edu_ultimo_anio_expansion': field(r['ultimo_anio_expansion'], '—'),
    }

# ── 6) cobertura + salida ─────────────────────────────────────────────
INDICADOR_IDS = [
    'edu_matricula_total', 'edu_var_matricula_2015_2025', 'edu_var_secundaria_2019_2025',
    'edu_alumnos_por_sede', 'edu_alumnos_por_seccion', 'edu_var_unidades_2015_2025',
    'edu_brecha_matricula_red', 'edu_abandono_secundario', 'edu_promocion_secundaria',
    'edu_sobreedad_secundaria', 'edu_participacion_estatal',
]
cobertura = {vid: sum(1 for m in municipios.values() if m[vid]['estado'] == 'disponible') for vid in INDICADOR_IDS}

out = {
    'meta': {
        'generado_por': 'scripts/build-monitor135-educacion-dataset.py',
        'fecha_actualizacion': '2026-08-14',
        'fuentes': {
            'consolidado_2025_2026': XLSX + ' — hoja RESUMEN_135',
            'serie_2015_2025': XLSX + ' — hoja EVOL_MUN_2015_2025',
            'maestro_territorial': MASTER_CSV,
        },
        'cobertura': {'municipios_totales': 135, **{vid + '_disponibles': n for vid, n in cobertura.items()}},
        'control_de_calidad': {
            'umbral_base_reducida_p10': base_p10,
            'limites_valor_extremo_iqr3': extreme_bounds,
            'anomalias_detectadas': anomaly_counts,
            'nota': 'Los valores marcados con quality_flag no se corrigen ni se ocultan; se excluyen únicamente del ranking cuando el flag es "valor_extremo" (ver monitor135-app.js). En ficha y comparador siempre son visibles.',
        },
    },
    'municipios': municipios,
}

with open(OUT_JSON, 'w', encoding='utf-8') as fh:
    json.dump(out, fh, ensure_ascii=False, indent=1)

print(f'{len(municipios)}/135 municipios procesados.')
print('Cobertura por indicador:', cobertura)
print('Anomalías detectadas:', anomaly_counts)
print('Escrito', OUT_JSON)
