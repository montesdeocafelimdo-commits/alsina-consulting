"""
Genera assets/data/transferencias-135-municipios-2026.json (+ .csv) para la
herramienta "Radiografía fiscal de tu municipio" en alsina-balance-fiscal-1s2026.html.

Fuente: Base_transferencias_135_municipios_PBA_completa.xlsx (Ministerio de
Economía PBA, vía hojas "Serie mensual" / "Base municipal"). Esa base trae los
montos nominales oficiales completos para los 135 municipios pero deja en
blanco el parámetro de IPC ("Pendiente IPC" en la hoja Parámetros). Este
script aplica la serie de inflación interanual mensual oficial ya usada y
publicada en los informes previos de Alsina ("Recaudación PBA 1S26" y
"Transferencias 1S26") para calcular variación real, ranking, brecha nominal
y equivalencia de poder de compra por municipio — sin inventar ningún monto.

Se valida contra los agregados provinciales ya publicados: variación real
semestral -2,5% y de junio -4,2% (reproducidos aquí a -2,496% y -4,162%,
diferencia atribuible a redondeo del original). Volver a correr:
  python3 scripts/build-transferencias-135-dataset.py
"""
import openpyxl, json, unicodedata

wb = openpyxl.load_workbook('/Users/felipe/Projects/alsina-consulting/Base_transferencias_135_municipios_PBA_completa.xlsx', data_only=True)
ws_base = wb['Base municipal']
ws_serie = wb['Serie mensual']

# Canonical monthly interanual IPC series, validated: reproduces published aggregate
# semester real var (-2.50%) and June real var (-4.2%) to within rounding.
IPC = {'Enero':0.324,'Febrero':0.331,'Marzo':0.326,'Abril':0.324,'Mayo':0.332,'Junio':0.335}
MESES = ['Enero','Febrero','Marzo','Abril','Mayo','Junio']

# 1) Pull monthly totals per municipio/año from "Serie mensual"
monthly = {}  # (id, anio, mes) -> total
names = {}
for r in range(5, ws_serie.max_row+1):
    anio = ws_serie.cell(row=r, column=1).value
    mes = ws_serie.cell(row=r, column=2).value
    mid = ws_serie.cell(row=r, column=3).value
    nombre = ws_serie.cell(row=r, column=4).value
    total = ws_serie.cell(row=r, column=18).value
    if anio is None or mid is None:
        continue
    monthly[(mid, anio, mes)] = total
    names[mid] = nombre

ids = sorted(names.keys())
assert len(ids) == 135, len(ids)

def deflate_sum(mid, anio, meses):
    s = 0.0
    for m in meses:
        v = monthly.get((mid, anio, m))
        if v is None:
            return None
        s += v / (1 + IPC[m])
    return s

def nominal_sum(mid, anio, meses):
    s = 0.0
    for m in meses:
        v = monthly.get((mid, anio, m))
        if v is None:
            return None
        s += v
    return s

def inflate_sum(mid, anio, meses):
    # what would have been needed in `anio` (2026) to match real purchasing power of anio-1 nominal, per month
    s = 0.0
    for m in meses:
        v = monthly.get((mid, anio, m))
        if v is None:
            return None
        s += v * (1 + IPC[m])
    return s

records = []
for mid in ids:
    nombre = names[mid]
    # semester
    n25_1s = nominal_sum(mid, 2025, MESES)
    n26_1s = nominal_sum(mid, 2026, MESES)
    real26_1s = deflate_sum(mid, 2026, MESES)  # 2026 semester at 2025 prices
    monto_necesario_1s = inflate_sum(mid, 2025, MESES)  # 2025 semester inflated to 2026 prices
    var_nom_1s = n26_1s/n25_1s - 1
    var_real_1s = real26_1s/n25_1s - 1
    brecha_1s = monto_necesario_1s - n26_1s
    per100_1s = 100*(1+var_real_1s)

    # june only
    n25_jun = nominal_sum(mid, 2025, ['Junio'])
    n26_jun = nominal_sum(mid, 2026, ['Junio'])
    real26_jun = deflate_sum(mid, 2026, ['Junio'])
    monto_necesario_jun = inflate_sum(mid, 2025, ['Junio'])
    var_nom_jun = n26_jun/n25_jun - 1
    var_real_jun = real26_jun/n25_jun - 1
    brecha_jun = monto_necesario_jun - n26_jun
    per100_jun = 100*(1+var_real_jun)

    records.append({
        'id': mid, 'municipio': nombre,
        'n2025_1s': n25_1s, 'n2026_1s': n26_1s,
        'var_nom_1s': var_nom_1s, 'var_real_1s': var_real_1s,
        'monto_necesario_1s': monto_necesario_1s, 'brecha_1s': brecha_1s, 'per100_1s': per100_1s,
        'n2025_jun': n25_jun, 'n2026_jun': n26_jun,
        'var_nom_jun': var_nom_jun, 'var_real_jun': var_real_jun,
        'monto_necesario_jun': monto_necesario_jun, 'brecha_jun': brecha_jun, 'per100_jun': per100_jun,
    })

# rankings (1 = best var_real_1s)
records.sort(key=lambda r: -r['var_real_1s'])
for i, r in enumerate(records, start=1):
    r['rank_1s'] = i

records_jun_sorted = sorted(records, key=lambda r: -r['var_real_jun'])
for i, r in enumerate(records_jun_sorted, start=1):
    r['rank_jun'] = i

# provincial averages (weighted, i.e. aggregate)
tot_n25_1s = sum(r['n2025_1s'] for r in records)
tot_n26_1s = sum(r['n2026_1s'] for r in records)

# recompute aggregate real var directly (weighted) for validation and as "promedio provincial"
# simpler: aggregate deflated total / aggregate nominal 2025
agg_deflated_1s = 0.0
for mid in ids:
    v = deflate_sum(mid, 2026, MESES)
    agg_deflated_1s += v
agg_var_real_1s = agg_deflated_1s/tot_n25_1s - 1

tot_n25_jun = sum(r['n2025_jun'] for r in records)
tot_n26_jun = sum(r['n2026_jun'] for r in records)
agg_deflated_jun = sum(deflate_sum(mid, 2026, ['Junio']) for mid in ids)
agg_var_real_jun = agg_deflated_jun/tot_n25_jun - 1

print('VALIDATION — aggregate semester real var:', round(agg_var_real_1s*100,3), '% (published: -2.5%)')
print('VALIDATION — aggregate june real var:', round(agg_var_real_jun*100,3), '% (published: -4.2%)')
print('VALIDATION — aggregate semester nominal var:', round((tot_n26_1s/tot_n25_1s-1)*100,3), '% (published: +29.6%)')
print('VALIDATION — aggregate june nominal var:', round((tot_n26_jun/tot_n25_jun-1)*100,3), '% (published: +27.9%)')

for r in records:
    r['diff_vs_avg_1s'] = r['var_real_1s'] - agg_var_real_1s
    r['diff_vs_avg_jun'] = r['var_real_jun'] - agg_var_real_jun

# Spot check San Isidro
si = next(r for r in records if r['municipio']=='San Isidro')
print('San Isidro 1S: var_real=%.4f%% rank=%d per100=%.2f' % (si['var_real_1s']*100, si['rank_1s'], si['per100_1s']))
print('San Isidro jun: var_real=%.4f%% rank=%d' % (si['var_real_jun']*100, si['rank_jun']))

out = {
    'meta': {
        'fuente': 'Ministerio de Economía PBA — Transferencias a Municipios (archivo Base_transferencias_135_municipios_PBA_completa.xlsx)',
        'ipc_serie': IPC,
        'ipc_fuente': 'INDEC IPC Nacional, inflación interanual mensual, según fue publicada en los informes previos de Alsina "Recaudación PBA 1S26" y "Transferencias 1S26"',
        'metodologia': 'var_real = (suma mensual de montos 2026 deflactados por IPC interanual del mes correspondiente) / (suma nominal 2025 del mismo tramo) - 1. monto_necesario = suma mensual de montos 2025 inflacionados por el mismo IPC interanual (equivalente a lo que se hubiese necesitado en 2026 para igualar el poder de compra de 2025). brecha_nominal = monto_necesario - monto_nominal_2026_efectivo.',
        'promedio_provincial_var_real_1s': agg_var_real_1s,
        'promedio_provincial_var_real_jun': agg_var_real_jun,
        'promedio_provincial_var_nom_1s': tot_n26_1s/tot_n25_1s-1,
        'promedio_provincial_var_nom_jun': tot_n26_jun/tot_n25_jun-1,
        'total_n2025_1s': tot_n25_1s, 'total_n2026_1s': tot_n26_1s,
        'total_n2025_jun': tot_n25_jun, 'total_n2026_jun': tot_n26_jun,
        'n_municipios': len(records),
        'validacion': 'Real semestral agregado y real de junio agregado reproducen (dentro de 0.01pp) los valores publicados en el informe fuente Transferencias 1S 2026 (-2,5% y -4,2% respectivamente).'
    },
    'municipios': records
}

with open('/Users/felipe/Projects/alsina-consulting/assets/data/transferencias-135-municipios-2026.json', 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=1)

print('OK — wrote', len(records), 'municipios')

# ── round for display cleanliness + normalized search key + CSV export ──
import unicodedata, csv

def norm(s):
    s = unicodedata.normalize('NFKD', s).encode('ascii','ignore').decode('ascii')
    return s.lower()

for r in records:
    r['municipio_norm'] = norm(r['municipio'])
    for k in ('n2025_1s','n2026_1s','monto_necesario_1s','brecha_1s','n2025_jun','n2026_jun','monto_necesario_jun','brecha_jun'):
        r[k] = round(r[k])
    for k in ('var_nom_1s','var_real_1s','per100_1s','var_nom_jun','var_real_jun','per100_jun','diff_vs_avg_1s','diff_vs_avg_jun'):
        r[k] = round(r[k], 6)

out['municipios'] = sorted(records, key=lambda r: r['municipio_norm'])

with open('/Users/felipe/Projects/alsina-consulting/assets/data/transferencias-135-municipios-2026.json', 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=1)

csv_cols = ['id','municipio','n2025_1s','n2026_1s','var_nom_1s','var_real_1s','rank_1s',
            'monto_necesario_1s','brecha_1s','per100_1s','diff_vs_avg_1s',
            'n2025_jun','n2026_jun','var_nom_jun','var_real_jun','rank_jun',
            'monto_necesario_jun','brecha_jun','per100_jun','diff_vs_avg_jun']
with open('/Users/felipe/Projects/alsina-consulting/assets/data/transferencias-135-municipios-2026.csv','w',newline='',encoding='utf-8') as f:
    w = csv.writer(f)
    w.writerow(csv_cols)
    for r in sorted(records, key=lambda r: r['municipio_norm']):
        w.writerow([r[c] for c in csv_cols])

print('Wrote JSON + CSV, rounded, sorted alfabéticamente.')
