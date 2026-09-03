"""
Fase 2 (corrección de datos) de Monitor 135.

Consolida y normaliza, para los 135 municipios de la Provincia de Buenos
Aires:

1. Identidad, gobierno, demografía y presupuesto — desde
   Info_Municipios_nueva/pba_municipios_master_v2.csv (fuente oficial con
   URL documentada por columna: SCBA, Censo 2022, Min. Economía PBA).
2. Transferencias por componente, primer semestre y junio, 2025 y 2026 —
   desde Base_transferencias_135_municipios_PBA_completa.xlsx, hoja
   "Serie mensual" (misma fuente y misma metodología de deflactación ya
   validada en scripts/build-transferencias-135-dataset.py contra los
   agregados provinciales publicados).

No inventa valores. Donde un campo no está disponible en la fuente
(presupuesto_2026, cud_pct_2026, intendente, etc.) queda marcado con
estado "no_disponible" en vez de completarse con 0.

Corrige además el bug de deflactor de municipios-data-hub.html
(`IPC_ANUAL = {"2024":2.178,"2025":2.17}`, un multiplicador anual
estimado aplicado al total del año completo): esa metodología NO se
reproduce acá. Los únicos cálculos reales de este archivo son
1S2025→1S2026 y jun2025→jun2026, con la serie de inflación interanual
mensual oficial ya documentada y validada. No hay datos mensuales
1S completos para 2023/2024 en las fuentes disponibles en este proyecto,
así que ese tramo NO se recalcula acá — ver metodología para el
tratamiento que se le da en la interfaz (nominal se conserva, "real" se
marca como pendiente de validación en vez de mostrar el número viejo).

Salida: private/data/monitor135-municipios.json

NOTA (FASE 5, AD-19): este dataset dejó de ser un archivo estático público
el 2026-08-21 — se sirve exclusivamente vía /api/monitor135/data.js,
filtrado por capacidad. Si en algún momento se vuelve a generar hacia
assets/data/, ese archivo queda servido sin ninguna protección de nuevo.
"""
import csv
import json
import unicodedata
import openpyxl

MASTER_CSV = 'Info_Municipios_nueva/pba_municipios_master_v2.csv'
TRANS_XLSX = 'Base_transferencias_135_municipios_PBA_completa.xlsx'
OUT_JSON = 'private/data/monitor135-municipios.json'

IPC = {'Enero': 0.324, 'Febrero': 0.331, 'Marzo': 0.326, 'Abril': 0.324, 'Mayo': 0.332, 'Junio': 0.335}
MESES = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio']

TRANS_COMPONENTES = [
    ('Coparticipación bruta', 'coparticipacion_bruta'),
    ('Omisión', 'omision'),
    ('Descentralización', 'descentralizacion'),
    ('Juegos de azar', 'juegos_de_azar'),
    ('FFPS', 'ffps_programas_sociales'),
    ('FSA', 'fsa_tratamiento_residuos'),
    ('Fort. recursos municipales', 'fortalecimiento_recursos_municipales'),
    ('Inclusión social', 'inclusion_social'),
    ('Financiamiento educativo', 'financiamiento_educativo'),
    ('Fortalecimiento fiscal', 'fortalecimiento_fiscal'),
    ('FEFIM', 'fefim'),
    ('Infraestructura 2017', 'infraestructura_2017'),
    ('Ley 14890', 'ley_14890'),
]

ALIAS_MASTER_TO_TRANS = {'MONTE': 'SAN MIGUEL DEL MONTE'}  # ya documentado en la hoja "Controles" del Excel


def norm(s):
    # Ñ/ñ es una letra propia, no una vocal acentuada: la protegemos de la
    # descomposición NFKD (que la partiría en "n" + tilde combinante) para
    # que coincida con las claves de MUNIS/GEO_DATA en municipios-data-hub.html
    # (ej.: "CAÑUELAS", no "CANUELAS").
    s = s.replace('ñ', '\x00').replace('Ñ', '\x00')
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii').upper().strip()
    return s.replace('\x00', 'Ñ')


def num_or_none(v, treat_zero_as_missing=True):
    """Convierte '', None o '0' (según corresponda) en None. No inventa un valor."""
    if v is None or v == '':
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    if treat_zero_as_missing and f == 0:
        return None
    return f


# ── 1) Identidad / gobierno / demografía / presupuesto ──────────────────
with open(MASTER_CSV, encoding='utf-8-sig') as f:
    master_rows = list(csv.DictReader(f))
assert len(master_rows) == 135, len(master_rows)

municipios = {}
for r in master_rows:
    key = norm(r['municipio_nombre_fuente'])
    presupuesto_2026 = num_or_none(r['presupuesto_2026'])
    pres_per_hab = num_or_none(r['presupuesto_per_capita_2026'])
    cud_pct = num_or_none(r['cud_pct_2026'], treat_zero_as_missing=False)  # 0% de CUD sería inusual pero no imposible; no forzamos missing
    intendente = r['intendente_actual_2025'] or None
    fuerza = r['fuerza_politica_2025'] or None

    municipios[key] = {
        'id': r['municipio_id_dataset'],
        'municipio': r['municipio_nombre'],
        'municipio_slug': r['municipio_slug'],
        'municipio_normalizado': key,
        'seccion_electoral': r['seccion_electoral'],
        'departamento_judicial': r['departamento_judicial'],
        'region': r['region_politica'],
        'es_conurbano': r['es_conurbano_24'] == 'True',
        'es_capital_provincial': r['es_capital_provincial'] == 'True',
        'poblacion': {
            'valor': num_or_none(r['poblacion_2022'], treat_zero_as_missing=False),
            'periodo': 2022,
            'fuente': 'Censo Nacional 2022 (INDEC) / SCBA',
            'url': r['fuente_censo_2022_general'] or r['fuente_poblacion_superficie_densidad'],
            'estado': 'disponible',
        },
        'superficie_km2': num_or_none(r['superficie_km2'], treat_zero_as_missing=False),
        'densidad_2022_hab_km2': num_or_none(r['densidad_2022_hab_km2'], treat_zero_as_missing=False),
        'categoria_poblacional_2022': r['categoria_poblacional_2022'],
        'gobierno': {
            'intendente_actual': {'valor': intendente, 'periodo': 2025, 'estado': 'disponible' if intendente else 'no_disponible'},
            'fuerza_politica': {'valor': fuerza, 'periodo': 2025, 'estado': 'disponible' if fuerza else 'no_disponible'},
            'electores_nacionales': num_or_none(r['electores_nacionales'], treat_zero_as_missing=False),
            'electores_extranjeros': num_or_none(r['electores_extranjeros'], treat_zero_as_missing=False),
        },
        'presupuesto_2026': {
            'valor': presupuesto_2026,
            'valor_per_capita': pres_per_hab,
            'periodo': 2026,
            'fuente': 'Ministerio de Economía PBA — Dirección de Finanzas Municipales',
            'url': r['finanzas_municipales_fuente'],
            'estado': 'disponible' if presupuesto_2026 is not None else 'no_publicado',
        },
        'cud_pct_2026': {
            'valor': cud_pct,
            'periodo': 2026,
            'fuente': 'Ministerio de Economía PBA',
            'estado': 'disponible' if cud_pct is not None else 'no_publicado',
        },
        'fecha_corte_fuente_gobierno_demografia': r['fecha_corte_dataset'],
    }

# ── 2) Transferencias por componente — 1S y junio, 2025 y 2026 ──────────
wb = openpyxl.load_workbook(TRANS_XLSX, data_only=True)
ws = wb['Serie mensual']
header = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
col_idx = {h: i + 1 for i, h in enumerate(header)}

monthly = {}  # (municipio_excel, anio, mes) -> {componente: valor}
for r in range(5, ws.max_row + 1):
    anio = ws.cell(row=r, column=col_idx['Año']).value
    mes = ws.cell(row=r, column=col_idx['Mes']).value
    muni = ws.cell(row=r, column=col_idx['Municipio']).value
    if anio is None or muni is None:
        continue
    row_vals = {comp_xlsx: (ws.cell(row=r, column=col_idx[comp_xlsx]).value or 0) for comp_xlsx, _ in TRANS_COMPONENTES}
    row_vals['Total'] = ws.cell(row=r, column=col_idx['Total']).value or 0
    monthly[(norm(muni), anio, mes)] = row_vals

trans_keys = set(k[0] for k in monthly.keys())


def resolve_trans_key(master_key):
    if master_key in trans_keys:
        return master_key
    if master_key in ALIAS_MASTER_TO_TRANS:
        return ALIAS_MASTER_TO_TRANS[master_key]
    return None


def sum_period(tkey, anio, meses, campo):
    return sum(monthly[(tkey, anio, m)][campo] for m in meses if (tkey, anio, m) in monthly)


def deflate_sum(tkey, anio, meses, campo):
    total = 0.0
    for m in meses:
        row = monthly.get((tkey, anio, m))
        if row is None:
            return None
        total += row[campo] / (1 + IPC[m])
    return total


def inflate_sum(tkey, anio, meses, campo):
    total = 0.0
    for m in meses:
        row = monthly.get((tkey, anio, m))
        if row is None:
            return None
        total += row[campo] * (1 + IPC[m])
    return total


matched, unmatched = 0, []
for key, m in municipios.items():
    tkey = resolve_trans_key(key)
    if tkey is None:
        unmatched.append(key)
        m['transferencias'] = None
        continue
    matched += 1

    def block(meses, anio_actual, anio_base):
        out = {'componentes_2026': {}, 'componentes_2025': {}}
        for comp_xlsx, comp_id in TRANS_COMPONENTES:
            out['componentes_2026'][comp_id] = round(sum_period(tkey, anio_actual, meses, comp_xlsx))
            out['componentes_2025'][comp_id] = round(sum_period(tkey, anio_base, meses, comp_xlsx))
        n25 = sum_period(tkey, anio_base, meses, 'Total')
        n26 = sum_period(tkey, anio_actual, meses, 'Total')
        real26 = deflate_sum(tkey, anio_actual, meses, 'Total')
        necesario = inflate_sum(tkey, anio_base, meses, 'Total')
        out['total_2025'] = round(n25)
        out['total_2026'] = round(n26)
        out['variacion_nominal'] = round(n26 / n25 - 1, 6) if n25 else None
        out['variacion_real'] = round(real26 / n25 - 1, 6) if (n25 and real26 is not None) else None
        out['monto_necesario_igualar_inflacion'] = round(necesario) if necesario is not None else None
        out['brecha_nominal'] = round(necesario - n26) if necesario is not None else None
        return out

    m['transferencias'] = {
        'primer_semestre': block(MESES, 2026, 2025),
        'junio': block(['Junio'], 2026, 2025),
        'fuente': 'Ministerio de Economía PBA — Transferencias a Municipios',
        'url': 'https://www.gba.gob.ar/economia/direccion_provincial_de_coordinacion_municipal_y_programas_de_desarrollo/transferencias_municipios',
        'metodologia_real': 'Deflactación mensual con IPC INDEC interanual (ver assets/data/monitor135-metodologia.json). No es el mismo método que el histórico 2023-2025 (deflactor anual) ya presente en el sitio: ese tramo no fue recalculado por falta de serie mensual oficial para esos años en este proyecto.',
        'cobertura_historica_previa': 'Los totales anuales 2023, 2024 y 2025 (serie preexistente en municipios-data-hub.html) se mantienen sin cambios en su valor nominal.',
    }

print(f'Transferencias: {matched}/135 municipios matcheados. Sin match: {unmatched}')

# ── rankings provinciales (excluye None; nunca asigna ranking a datos ausentes) ──
def add_rankings(periodo):
    rows = [(k, m['transferencias'][periodo]['variacion_real']) for k, m in municipios.items()
            if m['transferencias'] and m['transferencias'][periodo]['variacion_real'] is not None]
    rows.sort(key=lambda x: -x[1])
    for i, (k, _) in enumerate(rows, start=1):
        municipios[k]['transferencias'][periodo]['ranking_provincial'] = i
    for k, m in municipios.items():
        if m['transferencias'] and 'ranking_provincial' not in m['transferencias'][periodo]:
            m['transferencias'][periodo]['ranking_provincial'] = None
    total = sum(municipios[k]['transferencias'][periodo]['total_2026'] for k, _ in rows)
    base = sum(municipios[k]['transferencias'][periodo]['total_2025'] for k, _ in rows)
    deflated = sum(
        deflate_sum(resolve_trans_key(k), 2026, MESES if periodo == 'primer_semestre' else ['Junio'], 'Total')
        for k, _ in rows
    )
    return {'n': len(rows), 'promedio_provincial_var_real': round(deflated / base - 1, 6), 'promedio_provincial_var_nominal': round(total / base - 1, 6)}


meta_rank_1s = add_rankings('primer_semestre')
meta_rank_jun = add_rankings('junio')
print('Agregado 1S:', meta_rank_1s)
print('Agregado junio:', meta_rank_jun)

# ── cobertura ──
n_pres = sum(1 for m in municipios.values() if m['presupuesto_2026']['estado'] == 'disponible')
n_cud = sum(1 for m in municipios.values() if m['cud_pct_2026']['estado'] == 'disponible')
n_intendente = sum(1 for m in municipios.values() if m['gobierno']['intendente_actual']['estado'] == 'disponible')

out = {
    'meta': {
        'generado_por': 'scripts/build-monitor135-dataset.py',
        'fecha_actualizacion': '2026-09-02',
        'fuentes': {
            'identidad_gobierno_demografia_presupuesto': MASTER_CSV,
            'transferencias': TRANS_XLSX,
        },
        'ipc_serie_mensual_2026': IPC,
        'ipc_fuente': 'INDEC IPC Nacional, interanual mensual (misma serie validada en alsina-balance-fiscal-1s2026.html)',
        'formula_variacion_real': 'ver docstring de este script y assets/data/monitor135-metodologia.json',
        'promedio_provincial_1s': meta_rank_1s,
        'promedio_provincial_junio': meta_rank_jun,
        'cobertura': {
            'municipios_totales': 135,
            'municipios_con_transferencias_2026': matched,
            'municipios_con_presupuesto_2026': n_pres,
            'municipios_sin_presupuesto_2026': 135 - n_pres,
            'municipios_con_cud_2026': n_cud,
            'municipios_con_intendente_2025': n_intendente,
        },
    },
    'municipios': municipios,
}

with open(OUT_JSON, 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=1)
print('Escrito', OUT_JSON)
